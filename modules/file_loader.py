# modules/file_loader.py

import logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from typing import Tuple, Optional
import os

def load_file(file_path: str) -> Tuple[pd.DataFrame, Optional[Workbook]]:
    try:
        if file_path.endswith(".csv"):
            data = pd.read_csv(file_path)
            workbook = None  
            logging.info(f"Loaded CSV file: {file_path}")

        elif file_path.endswith((".xlsx", ".xls")):
            workbook = load_workbook(file_path, data_only=True)
            logging.info(f"Loaded Excel file: {file_path}")

            if "Master Asset List" in workbook.sheetnames:
                sheet = workbook["Master Asset List"]
                logging.info("Using 'Master Asset List' sheet.")
            else:
                sheet = workbook.active  
                logging.info(f"Using active sheet: {sheet.title}")

            data = pd.DataFrame(sheet.values)
            data.columns = data.iloc[0]  
            data = data[1:] 

            data.reset_index(drop=True, inplace=True)

        elif file_path.endswith(".json"):
            data = pd.read_json(file_path)
            workbook = None
            logging.info(f"Loaded JSON file: {file_path}")

        else:
            logging.error("Unsupported file format.")
            raise ValueError("Unsupported file format. Please provide a .csv, .xlsx, .xls, or .json file.")

        logging.info(f"File loaded successfully: {file_path}")
        return data, workbook

    except Exception as e:
        logging.exception("Error loading file.")
        raise e
