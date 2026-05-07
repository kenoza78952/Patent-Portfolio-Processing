# modules/formatter.py

import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
import pandas as pd

def apply_conditional_formatting(sheet, data: pd.DataFrame, formatting_rules: dict):
    try:
        cond_rules = formatting_rules.get("conditional_formatting", {}).get("rules", [])
        for rule in cond_rules:
            column = rule.get("column")
            condition = rule.get("condition")
            value = rule.get("value")
            fmt = rule.get("format", {})

            if column not in data.columns:
                logging.warning(f"Conditional formatting skipped. Column '{column}' not found.")
                continue

            col_idx = data.columns.get_loc(column) + 1
            column_letter = get_column_letter(col_idx)

            for row in range(2, len(data) + 2):
                cell = sheet[f"{column_letter}{row}"]
                cell_value = cell.value

                try:
                    if condition == "==":
                        condition_met = cell_value == value
                    elif condition == ">=":
                        try:
                            condition_met = float(cell_value) >= float(value)
                        except (TypeError, ValueError):
                            condition_met = False
                    elif condition == "<=":
                        try:
                            condition_met = float(cell_value) <= float(value)
                        except (TypeError, ValueError):
                            condition_met = False
                    else:
                        logging.warning(f"Unsupported condition '{condition}' in conditional formatting.")
                        condition_met = False

                    if condition_met:
                        if "font_color" in fmt or "bold" in fmt:
                            new_font = Font(
                                color=fmt.get("font_color", cell.font.color),
                                bold=fmt.get("bold", cell.font.bold)
                            )
                            cell.font = new_font
                except Exception as e:
                    logging.error(f"Error applying conditional formatting on cell {cell.coordinate}: {e}")

    except Exception as e:
        logging.exception("Error in apply_conditional_formatting function.")
        raise e

def apply_formatting(data: pd.DataFrame, workbook, innography_columns: dict, formatting_rules: dict) -> Workbook:
    try:
        if workbook is None:
            workbook = Workbook()
            sheet = workbook.active
            logging.info("Created a new workbook.")
        else:
            sheet = workbook.active
            logging.info(f"Using active sheet: {sheet.title}")

        sheet.title = "Master Asset List"

        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None

        header_style = formatting_rules["header"]
        header_fill_default = PatternFill(start_color=header_style["background_color"], fill_type="solid")
        client_header_fill = PatternFill(start_color="0874c4", fill_type="solid")  # Hex: #0874c4
        header_font_default = Font(
            name=header_style["font"]["name"],
            size=header_style["font"]["size"],
            bold=header_style["font"].get("bold", False)
        )
        header_alignment = Alignment(
            horizontal=header_style["alignment"]["horizontal"],
            vertical=header_style["alignment"]["vertical"],
            wrap_text=header_style.get("wrap_text", False)
        )
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        standard_columns = set(innography_columns.keys())

        headers = data.columns.tolist()
        for col_idx, column_name in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=column_name)
            
            if column_name in standard_columns or column_name == "Jurisdiction (US, Non-US)":
                if column_name == "Jurisdiction (US, Non-US)":
                    cell.fill = PatternFill(start_color="FF9494", fill_type="solid") 
                else:
                    cell.fill = header_fill_default
            else:
                cell.fill = client_header_fill 

            cell.font = header_font_default
            cell.alignment = header_alignment
            cell.border = thin_border
            sheet.column_dimensions[get_column_letter(col_idx)].width = header_style.get("column_width", 20)

        sheet.row_dimensions[1].height = header_style.get("row_height", 50)

        date_columns = ["Priority Date", "Filed Date", "Publish Date", "Est. Expiration Date"]

        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            sheet.row_dimensions[r_idx].height = formatting_rules["cell_format"].get("row_height", 50)
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                current_header = headers[c_idx - 1]

                if current_header == "Patent/ Publication Number":
                    if pd.notnull(value):
                        base_url = "https://patents.google.com/patent/"
                        hyperlink = f"{base_url}{str(value)}"
                        cell.hyperlink = hyperlink
                        cell.style = "Hyperlink"

                        cell.alignment = Alignment(
                            horizontal=formatting_rules["cell_format"]["alignment"]["horizontal"],
                            vertical=formatting_rules["cell_format"]["alignment"]["vertical"],
                            wrap_text=formatting_rules["cell_format"].get("wrap_text", False)
                        )

                        column_specific = formatting_rules.get("column_specific_formatting", {})
                        patent_format = column_specific.get("Patent/ Publication Number", {})
                        font_color = patent_format.get("font", {}).get("color", "0000FF")

                        cell.font = Font(
                            name=formatting_rules["cell_format"]["font"].get("name", "Century Gothic"),
                            size=formatting_rules["cell_format"]["font"].get("size", 11),
                            underline="single",
                            color=font_color
                        )

                if current_header in date_columns:
                    cell.number_format = 'MM/DD/YYYY'
                else:
                    if current_header != "Patent/ Publication Number":
                        cell.number_format = formatting_rules["cell_format"].get("number_format", "General")
                        
                if current_header != "Patent/ Publication Number":
                    cell.font = Font(
                        name=formatting_rules["cell_format"].get("font", {}).get("name", "Century Gothic"),
                        size=formatting_rules["cell_format"].get("font", {}).get("size", 11)
                    )
                    cell.alignment = Alignment(
                        horizontal=formatting_rules["cell_format"].get("alignment", {}).get("horizontal", "center"),
                        vertical=formatting_rules["cell_format"].get("alignment", {}).get("vertical", "center"),
                        wrap_text=formatting_rules["cell_format"].get("wrap_text", False)
                    )

        apply_conditional_formatting(sheet, data, formatting_rules)

        sheet.sheet_view.zoomScale = 80 

        logging.info("Formatting applied successfully.")
        return workbook
    except Exception as e:
        logging.error(f"Error applying formatting: {e}")
        raise
