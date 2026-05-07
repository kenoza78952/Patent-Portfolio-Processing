# overview_handler.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import sys
import os
import argparse
import logging

# =============================================================================
# STYLE PLACEHOLDERS AND HELPER FUNCTIONS
# =============================================================================

# Define a dictionary of style placeholders.
# Modify the properties below as needed.
STYLES = {
    "note": {
        "font": Font(color="FF0000", italic=True),
        "alignment": Alignment(horizontal="left"),
        "border": Border(),
        "fill": None
    },
    "data": {  # for non-table data cells (still thick borders, etc.)
        "font": Font(bold=True, color="000000"),
        "alignment": Alignment(horizontal="center"),
        "border": Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        ),
        "fill": None
    },
    "header": {  # for non-table header cells
        "font": Font(bold=True, color="000000"),
        "alignment": Alignment(horizontal="center"),
        "border": Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        ),
        "fill": PatternFill(start_color="FCFF33", end_color="FCFF33", fill_type="solid")
    },
    "default": {
        "font": Font(bold=True),
        "alignment": Alignment(horizontal="left"),
        "border": Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        ),
        "fill": None
    },
    # New styles for pivot tables:
    "table_header": {
        "font": Font(bold=True, color="000000"),
        "alignment": Alignment(horizontal="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        ),
        "fill": PatternFill(start_color="FCFF33", end_color="FCFF33", fill_type="solid")
    },
    "table_data": {
        "font": Font(bold=False, color="000000"),  # not bold
        "alignment": Alignment(horizontal="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        ),
        "fill": None
    }
}

def apply_style(cell, style_name):
    """
    Applies a style from the STYLES dictionary to the given cell.
    
    Available style names:
      - "note"    : For note cells (e.g., "*EUROPEAN Assets include ...", "Overview created ...")
      - "header"  : For table header cells (non-table headers)
      - "data"    : For all data value cells (non-table data)
      - "default" : For a default style
      - "table_header" : For pivot table header cells (thin border, bold)
      - "table_data"   : For pivot table data cells (thin border, not bold)
    """
    style = STYLES.get(style_name, STYLES["default"])
    cell.font = style["font"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]
    if style["fill"]:
        cell.fill = style["fill"]

def write_cell(sheet, cell_address, value, style_name="default"):
    """
    Writes a value to the cell at cell_address in the given sheet,
    and applies the specified style.
    
    Example usage:
       write_cell(sheet, "A1", "Overview created 02/03/2025", style_name="note")
       write_cell(sheet, "B2", 123, style_name="data")
       write_cell(sheet, "J1", "Jurisdiction Overview - All Assets", style_name="header")
    """
    cell = sheet[cell_address]
    cell.value = value
    apply_style(cell, style_name)

def add_totals(df, exclude_columns=None):
    """
    Adds a "Total" column to a DataFrame by summing all numeric columns (except the first column and any specified in exclude_columns).
    Automatically excludes any column whose name (or flattened name) includes 'year' (case-insensitive) from the summation.
    Then appends a totals row with "Grand Total" in the first column.
    Finally, renames the "Total" column to "Grand Total" so that the top-right cell of the table reads "Grand Total".
    
    Returns the resulting DataFrame.
    """
    if exclude_columns is None:
        exclude_columns = set()
    else:
        exclude_columns = set(exclude_columns)
    
    # Helper function to convert column name to a string (handles tuples for MultiIndex)
    def col_to_str(col):
        if isinstance(col, tuple):
            return " ".join(str(x) for x in col if x)
        return str(col)
    
    # Automatically exclude any column containing 'year' (case-insensitive)
    exclude_columns.update({col for col in df.columns if "year" in col_to_str(col).lower()})
    
    first_col = df.columns[0]
    # Only sum numeric columns beyond the first column and not in the exclude list.
    numeric_cols = [
        col for col in df.columns[1:]
        if pd.api.types.is_numeric_dtype(df[col]) and col not in exclude_columns
    ]
    df["Total"] = df[numeric_cols].sum(axis=1)
    
    # Create the totals row: force the first column to be "Grand Total"
    total_row = {}
    for col in df.columns:
        if col == first_col:
            total_row[col] = "Grand Total"
        elif pd.api.types.is_numeric_dtype(df[col]) and col not in exclude_columns:
            total_row[col] = df[col].sum()
        else:
            total_row[col] = ""
    df_total = pd.DataFrame([total_row])
    df = pd.concat([df, df_total], ignore_index=True)
    
    # Rename the totals column header from "Total" to "Grand Total"
    df = df.rename(columns={"Total": "Grand Total"})
    return df

def print_pivot_table(sheet, pivot_df, start_col, start_row, header_style="table_header", data_style="table_data"):
    """
    Prints a pivot table (DataFrame) into the sheet starting at the given column and row,
    using header_style for header cells and data_style for data cells.
    """
    # Write header row
    for i, header in enumerate(list(pivot_df.columns)):
        cell_address = f"{get_column_letter(start_col + i)}{start_row}"
        write_cell(sheet, cell_address, header, style_name=header_style)
    # Write data rows
    for r_idx, row_data in pivot_df.iterrows():
        for c_idx, value in enumerate(row_data):
            cell_address = f"{get_column_letter(start_col + c_idx)}{start_row + 1 + r_idx}"
            write_cell(sheet, cell_address, value, style_name=data_style)

# =============================================================================
# MAIN FUNCTION: create_overview
# =============================================================================
def create_overview(workbook_path: str):
    """
    Creates an Overview sheet in the given workbook with specific calculations,
    summary metrics (in Columns A, D, and G) and six summary tables.
    
    The first set of tables are for ALL assets/grants; then, after leaving three blank rows,
    the next set of tables are for ACTIVE assets only.
    
    The write_cell() function is used to write cells with styles.
    Pivot tables are printed using thin borders and non-bold data (only header rows are bold),
    and the totals row/column display "Grand Total" on both the bottom-left and top-right.
    """
    try:
        # Validate workbook path
        if not os.path.exists(workbook_path):
            print(f"Error: The workbook '{workbook_path}' does not exist.")
            return

        # Load the "Master Asset List" sheet into a DataFrame
        try:
            df = pd.read_excel(workbook_path, sheet_name="Master Asset List", engine='openpyxl')
            print("'Master Asset List' sheet loaded successfully.")
        except Exception as e:
            print("Error: Sheet 'Master Asset List' not found or cannot be read.")
            print(f"Details: {e}")
            return

        # Validate required columns
        required_columns = [
            "Type (Grant/Application)",
            "Publication Country",
            "Status (Active/Expired)",
            "Est. Expiration Date",
            "INPADOC Family ID",
            "Publish Date",
            "Filed Date",
            "Number of Forward Citations",
            "Patent/ Publication Number"
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"Error: Missing required columns in 'Master Asset List': {missing_columns}")
            return
        print("All required columns are present.")

        # Ensure date columns are datetime objects
        for col in ["Est. Expiration Date", "Publish Date", "Filed Date"]:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        if "Priority Date" in df.columns:
            df["Priority Date"] = pd.to_datetime(df["Priority Date"], errors='coerce')

        # Parameters and helper variables
        european_countries = {"DE", "GB", "FR", "EP"}
        target_date = datetime.strptime("02/03/2028", "%m/%d/%Y")
        today_str = datetime.today().strftime("%m/%d/%Y")

        # Standardize text for filtering
        df['Status Lower'] = df["Status (Active/Expired)"].str.lower()
        df['Type Lower'] = df["Type (Grant/Application)"].str.lower()

        # Create subsets using .copy() to avoid warnings
        active_df = df[df['Status Lower'] == "active"].copy()
        inactive_df = df[df['Status Lower'].isin(["inactive", "expired"])].copy()
        active_grants = active_df[active_df['Type Lower'] == "grant"].copy()
        inactive_grants = inactive_df[inactive_df['Type Lower'] == "grant"].copy()
        active_grants_3yr = active_grants[active_grants["Est. Expiration Date"] >= target_date].copy()

        # ----------------------------
        # COLUMN A: Overview Metrics (All Assets/Grants)
        # ----------------------------
        num_active_us_grants = active_grants[active_grants["Publication Country"].str.upper() == "US"].shape[0]
        num_active_cn_grants = active_grants[active_grants["Publication Country"].str.upper() == "CN"].shape[0]
        num_active_eu_grants = active_grants[active_grants["Publication Country"].str.upper().isin(european_countries)].shape[0]

        num_active_us_grants_3yr = active_grants_3yr[active_grants_3yr["Publication Country"].str.upper() == "US"].shape[0]
        num_active_cn_grants_3yr = active_grants_3yr[active_grants_3yr["Publication Country"].str.upper() == "CN"].shape[0]
        num_active_eu_grants_3yr = active_grants_3yr[active_grants_3yr["Publication Country"].str.upper().isin(european_countries)].shape[0]

        # ----------------------------
        # COLUMN D: Asset & Family Metrics
        # ----------------------------
        num_inpadoc_families = df["INPADOC Family ID"].nunique()
        num_active_assets = active_df[active_df["Patent/ Publication Number"].notna()].shape[0]
        num_inactive_assets = inactive_df[inactive_df["Patent/ Publication Number"].notna()].shape[0]
        num_provisional_applications = df[df['Status Lower'] == "provisional application"]
        num_provisional_applications = num_provisional_applications[num_provisional_applications["Patent/ Publication Number"].notna()].shape[0]
        num_unpublished_applications = df[df['Status Lower'] == "unpublished application"]
        num_unpublished_applications = num_unpublished_applications[num_unpublished_applications["Patent/ Publication Number"].notna()].shape[0]

        families_assets = active_df.groupby("INPADOC Family ID")["Publication Country"]\
                                  .apply(lambda x: set(x.str.upper())).reset_index(name="Countries")
        num_families_active_us_assets = families_assets["Countries"].apply(lambda s: "US" in s).sum()
        num_families_active_cn_assets = families_assets["Countries"].apply(lambda s: "CN" in s).sum()
        num_families_active_eu_assets = families_assets["Countries"].apply(lambda s: len(s.intersection(european_countries)) > 0).sum()
        num_families_active_us_cn_assets = families_assets["Countries"].apply(lambda s: ("US" in s and "CN" in s)).sum()
        num_families_active_us_eu_assets = families_assets["Countries"].apply(lambda s: ("US" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_families_active_cn_eu_assets = families_assets["Countries"].apply(lambda s: ("CN" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_families_active_us_cn_eu_assets = families_assets["Countries"].apply(lambda s: ("US" in s and "CN" in s and len(s.intersection(european_countries)) > 0)).sum()

        num_eu_assets = df[df["Publication Country"].str.upper().isin(european_countries)]["Patent/ Publication Number"].nunique()
        avg_forward_citations = df["Number of Forward Citations"].mean()

        # ----------------------------
        # COLUMN G: Grant & Family Metrics (Active Grants Only)
        # ----------------------------
        num_active_inpadoc_families = df[df["Status (Active/Expired)"].str.lower() == "active"]["INPADOC Family ID"].nunique()
        num_active_grants = active_grants[active_grants["Patent/ Publication Number"].notna()].shape[0]
        num_inactive_grants = inactive_grants[inactive_grants["Patent/ Publication Number"].notna()].shape[0]
        families_grants = active_grants.groupby("INPADOC Family ID")["Publication Country"]\
                                      .apply(lambda x: set(x.str.upper())).reset_index(name="Countries")
        num_fam_active_us_grants = families_grants["Countries"].apply(lambda s: "US" in s).sum()
        num_fam_active_cn_grants = families_grants["Countries"].apply(lambda s: "CN" in s).sum()
        num_fam_active_eu_grants = families_grants["Countries"].apply(lambda s: len(s.intersection(european_countries)) > 0).sum()
        num_fam_active_us_cn_grants = families_grants["Countries"].apply(lambda s: ("US" in s and "CN" in s)).sum()
        num_fam_active_us_eu_grants = families_grants["Countries"].apply(lambda s: ("US" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_fam_active_cn_eu_grants = families_grants["Countries"].apply(lambda s: ("CN" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_fam_active_us_cn_eu_grants = families_grants["Countries"].apply(lambda s: ("US" in s and "CN" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_eu_active_grants = active_grants[active_grants["Publication Country"].str.upper().isin(european_countries)]["Patent/ Publication Number"].nunique()

        families_grants_3yr = active_grants_3yr.groupby("INPADOC Family ID")["Publication Country"]\
                                              .apply(lambda x: set(x.str.upper())).reset_index(name="Countries")
        num_fam_active_us_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: "US" in s).sum()
        num_fam_active_cn_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: "CN" in s).sum()
        num_fam_active_eu_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: len(s.intersection(european_countries)) > 0).sum()
        num_fam_active_us_cn_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: ("US" in s and "CN" in s)).sum()
        num_fam_active_us_eu_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: ("US" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_fam_active_cn_eu_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: ("CN" in s and len(s.intersection(european_countries)) > 0)).sum()
        num_fam_active_us_cn_eu_grants_3yr = families_grants_3yr["Countries"].apply(lambda s: ("US" in s and "CN" in s and len(s.intersection(european_countries)) > 0)).sum()

        # ----------------------------
        # Prepare the Overview sheet
        # ----------------------------
        wb = load_workbook(workbook_path)
        if "Overview" in wb.sheetnames:
            del wb["Overview"]
            print("'Overview' sheet removed.")
        overview_sheet = wb.create_sheet(title="Overview")
        print("'Overview' sheet created.")

        # =============================================================================
        # WRITE OVERVIEW METRICS USING THE write_cell() HELPER WITH STYLE PLACEHOLDERS
        # =============================================================================

        # COLUMN A (Rows 1-5): All Active Grants
        write_cell(overview_sheet, "A1", f"Overview created {today_str}", style_name="note")
        write_cell(overview_sheet, "A2", "Number of Active US Grants", style_name="data")
        write_cell(overview_sheet, "B2", num_active_us_grants, style_name="data")
        write_cell(overview_sheet, "A3", "Number of Active CN Grants", style_name="data")
        write_cell(overview_sheet, "B3", num_active_cn_grants, style_name="data")
        write_cell(overview_sheet, "A4", "Number of Active EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "B4", num_active_eu_grants, style_name="data")
        # Note row (placeholder text; no adjacent calculation)
        write_cell(overview_sheet, "A5", "*EUROPEAN Assets include DE, GB, FR and EP Granted Patents and Applications", style_name="note")

        # COLUMN A (Rows 10-14): Grants with 3+ Years Remaining Life
        write_cell(overview_sheet, "A10", "Only Grants with 3+ Years of Remaining Life", style_name="note")
        write_cell(overview_sheet, "A11", "Number of Active US Grants", style_name="data")
        write_cell(overview_sheet, "B11", num_active_us_grants_3yr, style_name="data")
        write_cell(overview_sheet, "A12", "Number of Active CN Grants", style_name="data")
        write_cell(overview_sheet, "B12", num_active_cn_grants_3yr, style_name="data")
        write_cell(overview_sheet, "A13", "Number of Active EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "B13", num_active_eu_grants_3yr, style_name="data")
        # Note row (placeholder text)
        write_cell(overview_sheet, "A14", "*EUROPEAN Assets include DE, GB, FR and EP Granted Patents and Applications", style_name="note")

        # COLUMN D: Asset and Family Metrics
        write_cell(overview_sheet, "D2", "Number of INPADOC Families", style_name="data")
        write_cell(overview_sheet, "E2", num_inpadoc_families, style_name="data")
        write_cell(overview_sheet, "D4", "Number of Active Assets", style_name="data")
        write_cell(overview_sheet, "E4", num_active_assets, style_name="data")
        write_cell(overview_sheet, "D5", "Number of Inactive Assets", style_name="data")
        write_cell(overview_sheet, "E5", num_inactive_assets, style_name="data")
        write_cell(overview_sheet, "D6", "Number of Provisional Applications", style_name="data")
        write_cell(overview_sheet, "E6", num_provisional_applications, style_name="data")
        write_cell(overview_sheet, "D7", "Number of Unpublished Applications", style_name="data")
        write_cell(overview_sheet, "E7", num_unpublished_applications, style_name="data")
        write_cell(overview_sheet, "D10", "Number of Families with Active US Assets", style_name="data")
        write_cell(overview_sheet, "E10", num_families_active_us_assets, style_name="data")
        write_cell(overview_sheet, "D11", "Number of Families with Active CN Assets", style_name="data")
        write_cell(overview_sheet, "E11", num_families_active_cn_assets, style_name="data")
        write_cell(overview_sheet, "D12", "Number of Families with Active EUROPEAN Assets", style_name="data")
        write_cell(overview_sheet, "E12", num_families_active_eu_assets, style_name="data")
        write_cell(overview_sheet, "D13", "Number of Families with Active US and CN Assets", style_name="data")
        write_cell(overview_sheet, "E13", num_families_active_us_cn_assets, style_name="data")
        write_cell(overview_sheet, "D14", "Number of Families with Active US and EUROPEAN Assets", style_name="data")
        write_cell(overview_sheet, "E14", num_families_active_us_eu_assets, style_name="data")
        write_cell(overview_sheet, "D15", "Number of Families with Active CN and EUROPEAN Assets", style_name="data")
        write_cell(overview_sheet, "E15", num_families_active_cn_eu_assets, style_name="data")
        write_cell(overview_sheet, "D16", "Number of Families with Active US, CN and EUROPEAN Assets", style_name="data")
        write_cell(overview_sheet, "E16", num_families_active_us_cn_eu_assets, style_name="data")
        write_cell(overview_sheet, "D17", "*EUROPEAN Assets include DE, GB, FR and EP Granted Patents and Applications", style_name="note")
        write_cell(overview_sheet, "D20", "Average Number of Forward Citations", style_name="data")
        write_cell(overview_sheet, "E20", round(avg_forward_citations, 2) if not pd.isna(avg_forward_citations) else 0, style_name="data")
        write_cell(overview_sheet, "D21", "Number of Patents Involved in Litigations (Source: Innography)", style_name="data")
        write_cell(overview_sheet, "E21", 0, style_name="data")

        # COLUMN G: Grant and Family Metrics (Active Grants Only)
        write_cell(overview_sheet, "G2", "Number of Active INPADOC Families", style_name="data")
        write_cell(overview_sheet, "H2", df[df["Status (Active/Expired)"].str.lower() == "active"]["INPADOC Family ID"].nunique(), style_name="data")
        write_cell(overview_sheet, "G4", "Number of Active Grants", style_name="data")
        write_cell(overview_sheet, "H4", num_active_grants, style_name="data")
        write_cell(overview_sheet, "G5", "Number of Inactive Grants", style_name="data")
        write_cell(overview_sheet, "H5", num_inactive_grants, style_name="data")
        write_cell(overview_sheet, "G10", "Number of Families with Active US Grants", style_name="data")
        write_cell(overview_sheet, "H10", num_fam_active_us_grants, style_name="data")
        write_cell(overview_sheet, "G11", "Number of Families with Active CN Grants", style_name="data")
        write_cell(overview_sheet, "H11", num_fam_active_cn_grants, style_name="data")
        write_cell(overview_sheet, "G12", "Number of Families with Active EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H12", num_fam_active_eu_grants, style_name="data")
        write_cell(overview_sheet, "G13", "Number of Families with Active US and CN Grants", style_name="data")
        write_cell(overview_sheet, "H13", num_fam_active_us_cn_grants, style_name="data")
        write_cell(overview_sheet, "G14", "Number of Families with Active US and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H14", num_fam_active_us_eu_grants, style_name="data")
        write_cell(overview_sheet, "G15", "Number of Families with Active CN and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H15", num_fam_active_cn_eu_grants, style_name="data")
        write_cell(overview_sheet, "G16", "Number of Families with Active US, CN and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H16", num_fam_active_us_cn_eu_grants, style_name="data")
        write_cell(overview_sheet, "G17", "*EUROPEAN Assets include DE, GB, FR and EP Granted Patents", style_name="note")
        write_cell(overview_sheet, "G19", "Only Grants with 3+ Years of Remaining Life (Active after 02/03/2028)", style_name="note")
        write_cell(overview_sheet, "G20", "Number of Families with Active US Grants", style_name="data")
        write_cell(overview_sheet, "H20", num_fam_active_us_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G21", "Number of Families with Active CN Grants", style_name="data")
        write_cell(overview_sheet, "H21", num_fam_active_cn_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G22", "Number of Families with Active EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H22", num_fam_active_eu_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G23", "Number of Families with Active US and CN Grants", style_name="data")
        write_cell(overview_sheet, "H23", num_fam_active_us_cn_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G24", "Number of Families with Active US and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H24", num_fam_active_us_eu_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G25", "Number of Families with Active CN and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H25", num_fam_active_cn_eu_grants_3yr, style_name="data")
        write_cell(overview_sheet, "G26", "Number of Families with Active US, CN and EUROPEAN Grants", style_name="data")
        write_cell(overview_sheet, "H26", num_fam_active_us_cn_eu_grants_3yr, style_name="data")

        # =============================================================================
        # Create Summary Tables (Pivot-Style) for ALL Assets/Grants
        # =============================================================================
        if "Jurisdiction (US, Non-US)" not in df.columns:
            df["Jurisdiction (US, Non-US)"] = df["Publication Country"].apply(
                lambda x: "US" if str(x).strip().upper() == "US" else "Non-US"
            )
        # Table 1: Jurisdiction Overview – All Assets (Starting at cell J1)
        write_cell(overview_sheet, "J1", "Jurisdiction Overview - All Assets", style_name="note")
        pivot_juris = df.groupby(["Jurisdiction (US, Non-US)", "Type (Grant/Application)"])["Patent/ Publication Number"]\
                        .nunique().unstack(fill_value=0).reset_index()
        pivot_juris = add_totals(pivot_juris)
        start_col = 10  # Column J
        start_row = 2
        print_pivot_table(overview_sheet, pivot_juris, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_table1 = start_row + pivot_juris.shape[0]

        # Table 2: Priority Date Overview – All Assets (Starting at cell O1)
        priority_col_name = "Priority Date" if "Priority Date" in df.columns else "Publish Date"
        df[priority_col_name] = pd.to_datetime(df[priority_col_name], errors='coerce')
        df["Year"] = df[priority_col_name].dt.year
        write_cell(overview_sheet, "O1", "Priority Date Overview - All Assets", style_name="note")
        pivot_priority = df.groupby(["Year", "Jurisdiction (US, Non-US)", "Type (Grant/Application)"])["Patent/ Publication Number"]\
                           .nunique().reset_index()
        pivot_priority_table = pivot_priority.pivot_table(index="Year", columns=["Jurisdiction (US, Non-US)", "Type (Grant/Application)"],
                                                            values="Patent/ Publication Number", fill_value=0).reset_index()
        pivot_priority_table = add_totals(pivot_priority_table)
        # Flatten column names if MultiIndex
        pivot_priority_table.columns = [ " ".join(str(x) for x in col if x) if isinstance(col, tuple) else col for col in pivot_priority_table.columns ]
        start_col = 15  # Column O
        start_row = 2
        print_pivot_table(overview_sheet, pivot_priority_table, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_table2 = start_row + pivot_priority_table.shape[0]

        # Table 3: Expiration Date Overview – All Grants (Starting at cell X1)
        write_cell(overview_sheet, "X1", "Expiration Date Overview - All Grants", style_name="note")
        grants_df = df[df["Type (Grant/Application)"].str.lower() == "grant"].copy()
        grants_df["Expiration Year"] = grants_df["Est. Expiration Date"].dt.year
        pivot_exp = grants_df.groupby(["Expiration Year", "Jurisdiction (US, Non-US)"])["Patent/ Publication Number"]\
                             .nunique().reset_index()
        pivot_exp = pivot_exp.pivot_table(index="Expiration Year", columns="Jurisdiction (US, Non-US)",
                                          values="Patent/ Publication Number", fill_value=0).reset_index()
        pivot_exp = add_totals(pivot_exp)
        start_col = 24  # Column X
        start_row = 2
        print_pivot_table(overview_sheet, pivot_exp, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_table3 = start_row + pivot_exp.shape[0]

        # Determine the lowest ending row among the ALL Assets tables and leave 3 blank rows before ACTIVE tables
        max_table_end = max(end_row_table1, end_row_table2, end_row_table3)
        active_offset_row = max_table_end + 3

        # =============================================================================
        # Create Summary Tables for ACTIVE Assets Only
        # =============================================================================
        if "Jurisdiction (US, Non-US)" not in active_df.columns:
            active_df["Jurisdiction (US, Non-US)"] = active_df["Publication Country"].apply(
                lambda x: "US" if str(x).strip().upper() == "US" else "Non-US"
            )
        # Table 4: Jurisdiction Overview – Active Assets
        write_cell(overview_sheet, f"J{active_offset_row}", "Jurisdiction Overview - Active Assets", style_name="note")
        pivot_juris_active = active_df.groupby(["Jurisdiction (US, Non-US)", "Type (Grant/Application)"])["Patent/ Publication Number"]\
                                      .nunique().unstack(fill_value=0).reset_index()
        pivot_juris_active = add_totals(pivot_juris_active)
        start_col = 10  # Column J
        start_row = active_offset_row + 1
        print_pivot_table(overview_sheet, pivot_juris_active, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_active1 = active_offset_row + pivot_juris_active.shape[0] + 1

        # Table 5: Priority Date Overview – Active Assets
        write_cell(overview_sheet, f"O{active_offset_row}", "Priority Date Overview - Active Assets", style_name="note")
        active_df[priority_col_name] = pd.to_datetime(active_df[priority_col_name], errors='coerce')
        active_df["Year"] = active_df[priority_col_name].dt.year
        pivot_priority_active = active_df.groupby(["Year", "Jurisdiction (US, Non-US)", "Type (Grant/Application)"])["Patent/ Publication Number"]\
                                         .nunique().reset_index()
        pivot_priority_active_table = pivot_priority_active.pivot_table(index="Year", columns=["Jurisdiction (US, Non-US)", "Type (Grant/Application)"],
                                                                        values="Patent/ Publication Number", fill_value=0).reset_index()
        pivot_priority_active_table = add_totals(pivot_priority_active_table)
        pivot_priority_active_table.columns = [ " ".join(str(x) for x in col if x) if isinstance(col, tuple) else col for col in pivot_priority_active_table.columns ]
        start_col = 15  # Column O
        start_row = active_offset_row + 1
        print_pivot_table(overview_sheet, pivot_priority_active_table, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_active2 = start_row + pivot_priority_active_table.shape[0] + 1

        # Table 6: Expiration Date Overview – Active Grants
        write_cell(overview_sheet, f"X{active_offset_row}", "Expiration Date Overview - Active Grants", style_name="note")
        active_grants["Expiration Year"] = active_grants["Est. Expiration Date"].dt.year
        pivot_exp_active = active_grants.groupby(["Expiration Year", "Jurisdiction (US, Non-US)"])["Patent/ Publication Number"]\
                                         .nunique().reset_index()
        pivot_exp_active = pivot_exp_active.pivot_table(index="Expiration Year", columns="Jurisdiction (US, Non-US)",
                                                        values="Patent/ Publication Number", fill_value=0).reset_index()
        pivot_exp_active = add_totals(pivot_exp_active)
        start_col = 24  # Column X
        start_row = active_offset_row + 1
        print_pivot_table(overview_sheet, pivot_exp_active, start_col, start_row, header_style="table_header", data_style="table_data")
        end_row_active3 = start_row + pivot_exp_active.shape[0] + 1

        # =============================================================================
        # OPTIONAL: Apply extra borders to ranges if desired.
        # (Note: Our write_cell() calls already apply thick borders for non-table cells and thin borders for tables.)
        # =============================================================================
        def apply_borders(sheet, start_cell, end_cell):
            # This function is a placeholder if you want to reapply a different border style to a range.
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in sheet[start_cell:end_cell]:
                for cell in row:
                    cell.border = thin_border

        # =============================================================================
        # Adjust Column Widths (you can modify these values as needed)
        # =============================================================================
        column_widths = {
            'A': 40, 'B': 20, 'D': 60, 'E': 20, 'G': 60, 'H': 20,  
            'J': 30, 'K': 10, 'L': 10, 'M': 10, 'N': 10,
            'O': 30, 'P': 30, 'Q': 30, 'R': 30, 'S': 30, 'T': 30,
            'X': 30, 'Y': 20, 'Z': 20, 'AA': 20, 'AB': 20
        }
        for col, width in column_widths.items():
            overview_sheet.column_dimensions[col].width = width

        print(f"'Overview' sheet with summary tables created successfully in '{workbook_path}'.")
        wb.save(workbook_path)

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"Error: {e}")
        sys.exit(1)

# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    def parse_arguments():
        parser = argparse.ArgumentParser(
            description="Generate Overview Sheet with Summary Tables in Excel Workbook")
        parser.add_argument(
            'workbook_path',
            type=str,
            help='Path to the Excel workbook to modify.'
        )
        return parser.parse_args()

    args = parse_arguments()
    create_overview(args.workbook_path)
